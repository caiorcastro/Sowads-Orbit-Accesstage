# -*- coding: utf-8 -*-
"""
report_publicacao.py — Relatório MENSAL de publicação (base para cruzar com Google Search Console).

O que faz:
  1. Baixa o sitemap do blog do cliente (com <lastmod>).
  2. Diferencia do snapshot anterior -> quais URLs subiram desde o último report.
  3. Reconcilia o que NÓS geramos (lotes/temas) x o que está publicado (match por nome/Jaccard).
  4. Gera o report dos NOSSOS publicados (URL real + data + título + lote), pronto pro GSC.
  5. Atualiza o snapshot de dedup.

Uso:
  python3 tools/report_publicacao.py
  python3 tools/report_publicacao.py --gsc caminho/export_gsc_pages.csv   # já preenche clicks/impressões por URL

Saídas (output/reports/):
  publicados_report.csv/.xlsx     -> NOSSOS artigos publicados (para o GSC)
  reconciliacao_publicacao.csv    -> tudo (publicado + pendente) por lote
  urls_publicadas_accesstage.*    -> snapshot atualizado (dedup)
"""
import os, re, csv, sys, argparse, unicodedata, urllib.parse
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REP = os.path.join(BASE, "output/reports"); os.makedirs(REP, exist_ok=True)
SITEMAP = "https://blog.accesstage.com.br/sitemap.xml"

# Fontes do que geramos (lote, arquivo, coluna do título)
LOTES = [
    ("Lote 1", "output/articles/lote_veragi_temas.csv", "topic_pt"),
    ("Batch 2", "output/articles/lote_veragi_temas_batch2.csv", "topic_pt"),
    ("Batch 3", "output/articles/lote_veragi_temas_batch3.csv", "topic_pt"),
    ("Batch 4", "output/articles/lote_veragi_temas_batch4.csv", "topic_pt"),
    ("Lote 2", "output/articles/lote_veragi_lote2_combined.csv", "post_title"),
]
STOP = {"como","de","da","do","a","o","e","em","para","por","com","que","os","as","no","na","um","uma","sua","seu","x","sem","the","of","ja","nao"}

def ks(t):
    t = unicodedata.normalize("NFKD", str(t)).encode("ascii","ignore").decode().lower()
    return set(w for w in re.sub(r"[^a-z0-9 ]"," ",t).split() if len(w) > 3 and w not in STOP)

def fetch_sitemap(url):
    xml = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"}).text
    pub = []
    for b in re.findall(r"<url>(.*?)</url>", xml, flags=re.S):
        loc = re.search(r"<loc>([^<]+)</loc>", b); lm = re.search(r"<lastmod>([^<]+)</lastmod>", b)
        if loc:
            u = loc.group(1)
            slug = urllib.parse.unquote(u.rstrip("/").split("/")[-1])
            pub.append({"url": u, "date": lm.group(1)[:10] if lm else "", "keys": ks(slug)})
    return pub

def load_gsc(path):
    """Lê export do GSC (Pages). Retorna url -> {clicks, impressions, ctr, position}."""
    if not path or not os.path.exists(path): return {}
    rows = list(csv.reader(open(path, encoding="utf-8-sig")))
    if not rows: return {}
    head = [h.strip().lower() for h in rows[0]]
    def col(*names):
        for i, h in enumerate(head):
            if any(n in h for n in names): return i
        return -1
    iu, ic, ii, it, ip = col("url","página","pagina","page","top pages"), col("clic","click"), col("impress"), col("ctr"), col("posi","position")
    m = {}
    for r in rows[1:]:
        if iu < 0 or iu >= len(r): continue
        u = r[iu].strip()
        if not u.startswith("http"): continue
        g = lambda i: (r[i].strip() if 0 <= i < len(r) else "")
        m[u.rstrip("/")] = {"clicks": g(ic), "impr": g(ii), "ctr": g(it), "pos": g(ip)}
    return m

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sitemap", default=SITEMAP)
    ap.add_argument("--threshold", type=float, default=0.42)
    ap.add_argument("--gsc", default="", help="CSV de export do Google Search Console (Pages) para já preencher métricas")
    a = ap.parse_args()

    pub = fetch_sitemap(a.sitemap)
    print(f"Sitemap: {len(pub)} URLs (com data: {sum(1 for p in pub if p['date'])})")

    # diff vs snapshot anterior
    snap = os.path.join(REP, "urls_publicadas_accesstage.txt")
    prev = set(l.strip() for l in open(snap, encoding="utf-8")) if os.path.exists(snap) else set()
    now = set(p["url"] for p in pub)
    novos = sorted(now - prev)
    print(f"Novos desde o último snapshot: {len(novos)}")

    gsc = load_gsc(a.gsc)
    if a.gsc: print(f"GSC: {len(gsc)} páginas carregadas de {a.gsc}")

    # reconciliação
    recon = []; summ = {}
    for lote, path, col in LOTES:
        p = os.path.join(BASE, path)
        if not os.path.exists(p): continue
        rows = list(csv.DictReader(open(p, encoding="utf-8"))); n = 0
        for r in rows:
            t = (r.get(col) or "").strip()
            if not t: continue
            k = ks(t); best = (0.0, "", "")
            for pp in pub:
                j = len(k & pp["keys"]) / max(1, len(k | pp["keys"]))
                if j > best[0]: best = (j, pp["url"], pp["date"])
            ok = best[0] >= a.threshold
            if ok: n += 1
            recon.append({"lote": lote, "tema": t, "status": "PUBLICADO" if ok else "pendente",
                          "sim": round(best[0], 2), "url": best[1] if ok else "", "data": best[2] if ok else ""})
        summ[lote] = (len(rows), n)

    with open(os.path.join(REP, "reconciliacao_publicacao.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["lote","tema","status","sim","url","data"]); w.writeheader(); w.writerows(recon)

    # report dos NOSSOS publicados (URLs únicas, melhor match)
    best_by_url = {}
    for r in recon:
        if r["status"] != "PUBLICADO" or not r["url"]: continue
        u = r["url"]
        if u not in best_by_url or r["sim"] > best_by_url[u]["sim"]: best_by_url[u] = r
    data = sorted(best_by_url.values(), key=lambda r: (r["data"], r["sim"]), reverse=True)

    with open(os.path.join(REP, "publicados_report.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(["data_lastmod","url","titulo_nosso","lote","confianca","clicks","impressoes","ctr","posicao"])
        for r in data:
            g = gsc.get(r["url"].rstrip("/"), {})
            w.writerow([r["data"], r["url"], r["tema"], r["lote"], r["sim"], g.get("clicks",""), g.get("impr",""), g.get("ctr",""), g.get("pos","")])

    # xlsx
    wb = Workbook(); ws = wb.active; ws.title = "Publicados"
    NAVY = "1F3A5F"; thin = Side(style="thin", color="C7D2DD"); bd = Border(left=thin,right=thin,top=thin,bottom=thin)
    head = ["Data","URL publicada","Título (nosso)","Lote","Conf.","Clicks","Impressões","CTR","Posição"]
    ws.merge_cells("A1:I1"); c = ws["A1"]
    c.value = f"Accesstage — NOSSOS artigos publicados ({len(data)}) · cruzar com Google Search Console"
    c.font = Font(size=13, bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(vertical="center", indent=1); ws.row_dimensions[1].height = 26
    for j, h in enumerate(head, 1):
        c = ws.cell(row=2, column=j, value=h); c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="16293F"); c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bd
    for i, r in enumerate(data, 3):
        g = gsc.get(r["url"].rstrip("/"), {}); z = "F4F7FA" if i % 2 else "FFFFFF"; low = r["sim"] < 0.5
        vals = [r["data"], r["url"], r["tema"], r["lote"], ("⚠ "+str(r["sim"])) if low else r["sim"],
                g.get("clicks",""), g.get("impr",""), g.get("ctr",""), g.get("pos","")]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v); c.border = bd; c.fill = PatternFill("solid", fgColor=z)
            c.font = Font(size=10, color="C0392B" if (low and j == 5) else "222222")
            c.alignment = Alignment(vertical="center", indent=1, wrap_text=(j == 3))
    for col, wd in {"A":11,"B":58,"C":46,"D":9,"E":8,"F":9,"G":11,"H":8,"I":9}.items(): ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A3"; ws.auto_filter.ref = "A2:I2"
    wb.save(os.path.join(REP, "publicados_report.xlsx"))

    # atualiza snapshot de dedup
    urls = sorted(now)
    with open(os.path.join(REP, "urls_publicadas_accesstage.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f); w.writerow(["url","titulo_aprox"])
        for u in urls:
            w.writerow([u, urllib.parse.unquote(u.rstrip("/").split("/")[-1]).replace("-"," ").capitalize()])
    open(os.path.join(REP, "urls_publicadas_accesstage.txt"), "w", encoding="utf-8").write("\n".join(urls) + "\n")

    print("\n== RECONCILIAÇÃO (gerado x publicado) ==")
    tg = tp = 0
    for lote, (gtot, p) in summ.items():
        print(f"  {lote:9} gerados {gtot:3} | publicados {p:3} | pendentes {gtot-p:3}"); tg += gtot; tp += p
    print(f"  {'TOTAL':9} gerados {tg:3} | publicados {tp:3} | pendentes {tg-tp:3}")
    print(f"\n✓ report: output/reports/publicados_report.xlsx (+ .csv) | {len(data)} publicados nossos")
    print(f"  reconciliação: output/reports/reconciliacao_publicacao.csv | snapshot atualizado: {len(urls)} URLs")

if __name__ == "__main__":
    main()
