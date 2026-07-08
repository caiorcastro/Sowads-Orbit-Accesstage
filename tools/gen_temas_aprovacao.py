# -*- coding: utf-8 -*-
"""
gen_temas_aprovacao.py — Gera N temas NOVOS (deduplicados) com metadados de aprovação
e monta um xlsx diagramado (igual ao Lote 2) para o cliente aprovar.

Uso: python3 tools/gen_temas_aprovacao.py <lote_num> [qtd]     (ex.: 3 20)
Saída: output/reports/temas_aprovacao_lote<N>.xlsx + content_plan/temas_lote<N>.csv
"""
import os, csv, json, re, sys, unicodedata, requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
load_dotenv(os.path.join(BASE, ".env"))
KEY = os.getenv("OPENROUTER_API_KEY")
URL = "https://openrouter.ai/api/v1/chat/completions"
MODEL = "google/gemini-2.5-flash"
LOTE = int(sys.argv[1]) if len(sys.argv) > 1 else 3
TARGET = int(sys.argv[2]) if len(sys.argv) > 2 else 20
CATS = ["SEO & AIO", "Estratégia e Performance", "Data e Analytics"]

def norm(t):
    t = unicodedata.normalize("NFKD", str(t)).encode("ascii","ignore").decode().lower()
    return re.sub(r"[^a-z0-9 ]", " ", t)
def keyset(t):
    stop = {"como","de","da","do","a","o","e","em","para","por","com","que","os","as","no","na","um","uma","sua","seu","the","of","x"}
    return set(w for w in norm(t).split() if len(w) > 3 and w not in stop)

existing = []
u = os.path.join(BASE, "output/reports/urls_publicadas_accesstage.csv")
if os.path.exists(u):
    for r in csv.DictReader(open(u, encoding="utf-8")): existing.append(r.get("titulo_aprox",""))
for f in ["lote_veragi_temas.csv","lote_veragi_temas_batch2.csv","lote_veragi_temas_batch3.csv","lote_veragi_temas_batch4.csv","lote_veragi_lote2_combined.csv"]:
    p = os.path.join(BASE, "output/articles", f)
    if os.path.exists(p):
        for r in csv.DictReader(open(p, encoding="utf-8")): existing.append(r.get("topic_pt") or r.get("post_title") or "")
cp = os.path.join(BASE, "content_plan")
if os.path.isdir(cp):
    for f in os.listdir(cp):
        if f.startswith("temas") and f.endswith(".csv"):
            for r in csv.DictReader(open(os.path.join(cp, f), encoding="utf-8")): existing.append(r.get("topic_pt",""))
existing_keys = [keyset(t) for t in existing if t and t.strip()]
print(f"Dedup base: {len(existing_keys)} títulos")

def is_dup(title, acc):
    k = keyset(title)
    if not k: return True
    for ek in existing_keys + acc:
        if ek and len(k & ek)/max(1,len(k|ek)) >= 0.5: return True
    return False

MODULES = ("Contas a Pagar; Contas a Receber; Tesouraria; Crédito/Risco Sacado/Antecipação; Analytics/Dados; "
           "Integrações Bancárias (EDI, API, Open Finance, CNAB, VAN); Cash Pooling; Estratégia/Governança/Compliance")

def parse_objs(txt):
    out = []
    for m in re.finditer(r"\{[^{}]*\}", txt, re.S):
        for c in (m.group(0), m.group(0).replace("'", '"')):
            try: out.append(json.loads(c)); break
            except Exception: continue
    return out

def gen(nreq, avoid):
    usr = f"""Estrategista de conteúdo SEO B2B da Accesstage (fintech; produto Veragi — gestão financeira corporativa). Público: CFOs, tesoureiros, controllers de empresas médias/grandes no Brasil.

Gere {nreq} temas de artigo NOVOS, em português-BR, específicos e acionáveis, pelos módulos: {MODULES}.
NÃO repita nem seja parecido com: {"; ".join(avoid[-140:])}

Para CADA tema devolva um objeto JSON com:
- "topic_pt": título do artigo
- "funnel": "Topo" | "Meio" | "Fundo"
- "module": módulo Veragi
- "keyword": keyword principal em pt-br (2-4 palavras)
- "volume": faixa estimada de buscas/mês no Brasil, uma de ["10–50","50–200","200–500","500–1k","1k–5k"]
- "rating": inteiro 1-5 (potencial = volume x intenção)
- "defesa": 1 frase curta defendendo por que vale publicar (gancho estratégico)

Responda SÓ com um array JSON de {nreq} objetos."""
    try:
        r = requests.post(URL, headers={"Authorization": f"Bearer {KEY}"},
            json={"model": MODEL, "messages": [{"role":"user","content":usr}], "temperature": 0.9, "max_tokens": 3200}, timeout=100)
        return parse_objs(r.json()["choices"][0]["message"]["content"])
    except Exception as e:
        print("  (gen erro:", str(e)[:60], ")"); return []

got, acc, avoid = [], [], [t for t in existing if t and t.strip()]
tries = 0
while len(got) < TARGET and tries < 8:
    tries += 1
    for it in gen(26, avoid):
        title = (it.get("topic_pt") or "").strip()
        if not title or is_dup(title, acc): continue
        it["category"] = "SEO & AIO"
        try: it["rating"] = str(int(it.get("rating", 3)))
        except: it["rating"] = "3"
        got.append(it); acc.append(keyset(title)); avoid.append(title)
        if len(got) >= TARGET: break
    print(f"  tentativa {tries}: {len(got)}/{TARGET}")
got = got[:TARGET]

# ordena por funil (Topo, Meio, Fundo)
order = {"Topo":0,"Meio":1,"Fundo":2}
got.sort(key=lambda x: order.get(x.get("funnel","Meio"), 1))

# CSV p/ o content_engine
os.makedirs(cp, exist_ok=True)
with open(os.path.join(cp, f"temas_lote{LOTE}.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f); w.writerow(["topic_pt","vertical","category"])
    for t in got: w.writerow([t["topic_pt"], "fintech", "SEO & AIO"])

# ── xlsx diagramado (mesmo estilo do Lote 2) ──
NAVY="1F3A5F"; NAVY_D="16293F"; GREY="5B6B7B"; WHITE="FFFFFF"; ZEBRA="F4F7FA"
BLUE="DCE9F5"; AMBER="FCEFCF"; GREEN="DBEFDD"
FUNIL_FILL={"Topo":BLUE,"Meio":AMBER,"Fundo":GREEN}; FUNIL_TXT={"Topo":"1F4E79","Meio":"8A5A00","Fundo":"1E6B2E"}
RATING_FILL={"5":"2E7D32","4":"66A86B","3":"E0A800","2":"E8833A","1":"C0392B"}
thin=Side(style="thin",color="C7D2DD"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
wb=Workbook(); ws=wb.active; ws.title=f"Temas Lote {LOTE} — Aprovação"
ws.merge_cells("A1:G1"); c=ws["A1"]; c.value=f"Plataforma Veragi · Accesstage — {len(got)} Temas para Aprovação (Lote {LOTE})"
c.font=Font(size=16,bold=True,color=WHITE); c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=Alignment(vertical="center",indent=1); ws.row_dimensions[1].height=30
ws.merge_cells("A2:G2"); c=ws["A2"]
c.value=("Volume = ESTIMATIVA de buscas/mês no Brasil (trocar pelo Google Ads)   ·   Rating 1–5 = potencial (volume × intenção)   ·   "
         "Funil: Topo=descoberta  Meio=consideração  Fundo=decisão   ·   Deduplicado vs. 385 URLs publicadas + lotes anteriores")
c.font=Font(size=9,italic=True,color=WHITE); c.fill=PatternFill("solid",fgColor=GREY); c.alignment=Alignment(vertical="center",indent=1); ws.row_dimensions[2].height=24
headers=["#","Tema (título do artigo)","Funil","Módulo Veragi","Keyword principal","Vol/mês (est.)","★"]
for j,h in enumerate(headers,1):
    c=ws.cell(row=3,column=j,value=h); c.font=Font(bold=True,color=WHITE,size=10); c.fill=PatternFill("solid",fgColor=NAVY_D)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=border
ws.row_dimensions[3].height=22
r=4
for i,t in enumerate(got,1):
    fun=t.get("funnel","Meio"); rat=t.get("rating","3"); zebra=ZEBRA if i%2==0 else WHITE
    vals=[i,t["topic_pt"],fun,t.get("module","")[:22],t.get("keyword",""),t.get("volume",""),rat]
    for j,v in enumerate(vals,1):
        c=ws.cell(row=r,column=j,value=v); c.border=border; c.fill=PatternFill("solid",fgColor=zebra)
        if j==1: c.alignment=Alignment(horizontal="center",vertical="center"); c.font=Font(size=9,color=GREY)
        elif j==2: c.alignment=Alignment(vertical="center",wrap_text=True,indent=1); c.font=Font(size=10,bold=True,color="1A1A1A")
        elif j==3: c.fill=PatternFill("solid",fgColor=FUNIL_FILL.get(fun,BLUE)); c.font=Font(size=9,bold=True,color=FUNIL_TXT.get(fun,"1F4E79")); c.alignment=Alignment(horizontal="center",vertical="center")
        elif j==7: c.fill=PatternFill("solid",fgColor=RATING_FILL.get(rat,"E0A800")); c.font=Font(size=11,bold=True,color=WHITE); c.alignment=Alignment(horizontal="center",vertical="center")
        else: c.alignment=Alignment(vertical="center",wrap_text=True,indent=1); c.font=Font(size=9,color="333333")
    r+=1
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=7)
    c=ws.cell(row=r,column=2,value="↳ Defesa:  "+t.get("defesa","")); c.font=Font(size=8,italic=True,color=GREY)
    c.alignment=Alignment(vertical="center",wrap_text=True,indent=1); c.fill=PatternFill("solid",fgColor=zebra)
    ws.cell(row=r,column=1).fill=PatternFill("solid",fgColor=zebra); ws.cell(row=r,column=1).border=border
    for j in range(2,8): ws.cell(row=r,column=j).border=border
    ws.row_dimensions[r].height=22; r+=1
for col,wd in {"A":4,"B":54,"C":9,"D":16,"E":26,"F":14,"G":5}.items(): ws.column_dimensions[col].width=wd
ws.freeze_panes="A4"; ws.auto_filter.ref="A3:G3"
ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
ws.sheet_properties.pageSetUpPr.fitToPage=True; ws.print_title_rows="1:3"
out=os.path.join(BASE,"output/reports",f"temas_aprovacao_lote{LOTE}.xlsx"); wb.save(out)
from collections import Counter
print(f"\n✓ {len(got)} temas -> {out}")
print("Funil:", dict(Counter(t.get('funnel','') for t in got)))
for t in got: print(f"  [{t.get('funnel','')[:4]:4}] {t['topic_pt']}")
